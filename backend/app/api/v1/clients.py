"""Clients API - CRUD for clients and their accounts."""
import io
import uuid
import random
from datetime import date, datetime

import openpyxl
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from sqlalchemy import update as sa_update
from app.db.session import get_db
from app.core.deps import CurrentUser, get_current_user
from app.core.config import settings
from app.models.client import Client, ClientAccount
from app.schemas.client import (
    ClientCreate,
    ClientUpdate,
    ClientResponse,
    AccountCreate,
    AccountUpdate,
    AccountResponse,
)
from app.schemas.client_portal import ClientPortalUpdate
from app.services import client_service
from app.services import client_portal_service
from app.services.email_service import send_portal_link_email

router = APIRouter(prefix="/clients", tags=["clients"])


async def _gen_unique_code(db: AsyncSession) -> str:
    """중복 없는 6자리 고유번호 생성."""
    while True:
        code = str(random.randint(100000, 999999))
        result = await db.execute(select(Client).where(Client.unique_code == code))
        if not result.scalar_one_or_none():
            return code


def _parse_birth_date(value) -> date | None:
    """엑셀 datetime 객체 또는 YYYY-MM-DD 문자열을 date로 변환."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """엑셀 파일로 고객 대량 등록.

    템플릿 컬럼: [No., 고객명, 고유번호, 생년월일, 전화번호, 이메일]
    No.와 고유번호는 무시합니다 (자동 생성).
    1행은 헤더로 무시합니다.
    """
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="유효하지 않은 엑셀 파일입니다.")

    ws = wb.active
    created = 0
    skipped = 0
    skipped_names: list[str] = []
    errors: list[str] = []

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row_idx, row in enumerate(rows, start=2):
        # 완전히 빈 행은 무시
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        # 컬럼 매핑: [고객명, 생년월일, 전화번호, 이메일]
        name = str(row[0]).strip() if row[0] is not None else ""
        birth_raw = row[1] if len(row) > 1 else None
        phone_raw = row[2] if len(row) > 2 else None
        email_raw = row[3] if len(row) > 3 else None

        if not name:
            errors.append(f"{row_idx}행: 고객명 누락")
            continue

        phone_str = str(phone_raw).strip() if phone_raw is not None else None
        # 엑셀에서 숫자로 읽힌 전화번호 처리 (예: 1012345678.0 → 01012345678)
        if phone_str and phone_str.endswith(".0"):
            phone_str = phone_str[:-2]
        email_str = str(email_raw).strip() if email_raw is not None else None
        birth_date = _parse_birth_date(birth_raw)

        # 중복 체크 (이름 + 전화번호)
        dup_query = select(Client).where(
            Client.user_id == current_user.id,
            Client.name == name,
        )
        if phone_str:
            dup_query = dup_query.where(Client.phone == phone_str)
        dup_result = await db.execute(dup_query)
        if dup_result.scalar_one_or_none():
            skipped += 1
            skipped_names.append(f"{row_idx}행: {name}")
            continue

        client = Client(
            id=str(uuid.uuid4()),
            user_id=current_user.id,
            name=name,
            unique_code=await _gen_unique_code(db),
            birth_date=birth_date,
            phone=phone_str or None,
            email=email_str or None,
        )
        db.add(client)
        created += 1

    await db.commit()
    return {"created": created, "skipped": skipped, "skipped_names": skipped_names, "errors": errors}


@router.get("/template-excel")
async def download_template(
    token: Optional[str] = Query(None),
    current_user=Depends(get_current_user),
):
    """빈 엑셀 템플릿 다운로드 (업로드용)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "고객등록"
    headers = ["고객명", "생년월일", "전화번호", "이메일"]
    ws.append(headers)
    # 예시 행
    ws.append(["홍길동", "1990-01-15", "010-1234-5678", "hong@example.com"])
    # 헤더 스타일
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 25
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_template.xlsx"},
    )


@router.get("/download-excel")
async def download_excel(
    token: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """현재 사용자의 고객 목록을 엑셀 파일로 다운로드.

    window.open 방식으로 호출되므로 쿼리 파라미터 token으로 인증합니다.
    """
    from jose import jwt, JWTError
    from app.models.user import User

    if not token:
        raise HTTPException(status_code=401, detail="Token required")
    try:
        payload = jwt.decode(token, settings.SECRET_KEY, algorithms=["HS256"])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

    result = await db.execute(
        select(Client)
        .where(Client.user_id == user_id)
        .order_by(Client.created_at)
    )
    clients = result.scalars().all()

    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "고객정보"
    headers = ["No.", "고객명", "고유번호", "생년월일", "전화번호", "이메일"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, c in enumerate(clients, 1):
        ws.append([
            idx,
            c.name,
            c.unique_code or "",
            str(c.birth_date) if c.birth_date else "",
            c.phone or "",
            c.email or "",
        ])

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customers.xlsx"},
    )


@router.get("", response_model=list[ClientResponse])
async def list_clients(
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await client_service.list_clients(db, current_user.id)


@router.post("", response_model=ClientResponse, status_code=201)
async def create_client(
    body: ClientCreate,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await client_service.create_client(
        db, current_user.id, body.name, body.memo, body.ssn
    )


@router.get("/{client_id}", response_model=ClientResponse)
async def get_client(
    client_id: str,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    client = await client_service.get_client(db, current_user.id, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    return client


@router.put("/{client_id}", response_model=ClientResponse)
async def update_client(
    client_id: str,
    body: ClientUpdate,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    client = await client_service.update_client(
        db, current_user.id, client_id, body.name, body.memo, body.ssn
    )
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    return client


@router.delete("/{client_id}", status_code=204)
async def delete_client(
    client_id: str,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ok = await client_service.delete_client(db, current_user.id, client_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Client not found")


@router.get("/{client_id}/accounts", response_model=list[AccountResponse])
async def list_accounts(
    client_id: str,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    client = await client_service.get_client(db, current_user.id, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    return await client_service.list_accounts(db, client_id)


@router.post("/{client_id}/accounts", response_model=AccountResponse, status_code=201)
async def create_account(
    client_id: str,
    body: AccountCreate,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    client = await client_service.get_client(db, current_user.id, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    return await client_service.create_account(
        db,
        client_id,
        body.account_type,
        body.account_number,
        body.securities_company,
        body.representative,
        body.monthly_payment,
    )


@router.put("/{client_id}/accounts/{account_id}", response_model=AccountResponse)
async def update_account(
    client_id: str,
    account_id: str,
    body: AccountUpdate,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    account = await client_service.update_account(
        db, account_id, client_id, body.model_dump(exclude_unset=True)
    )
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    return account


@router.delete("/{client_id}/accounts/{account_id}", status_code=204)
async def delete_account(
    client_id: str,
    account_id: str,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ok = await client_service.delete_account(db, account_id, client_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Account not found")


@router.post("/{client_id}/send-portal-link")
async def send_portal_link(
    client_id: str,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return portal view link and optionally send it by email.

    Response:
    - portal_link: the full URL to the client portal
    - email_sent: True if email was actually dispatched
    - message: human-readable status
    """
    client = await client_service.get_client(db, current_user.id, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    if not client.portal_token:
        raise HTTPException(status_code=400, detail="Client has no portal token")

    portal_link = f"{settings.FRONTEND_URL}/client/{client.portal_token}"

    email_sent = False
    if client.email:
        email_sent = await send_portal_link_email(
            client_email=client.email,
            client_name=client.name,
            portal_link=portal_link,
        )
        message = "링크가 이메일로 발송되었습니다." if email_sent else "링크를 생성했습니다. (이메일 미발송)"
    else:
        message = "이메일이 등록되지 않아 링크만 반환합니다."

    return {
        "portal_link": portal_link,
        "email_sent": email_sent,
        "message": message,
    }


@router.patch("/{client_id}", response_model=ClientResponse)
async def patch_client_portal_info(
    client_id: str,
    body: ClientPortalUpdate,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update portal-related fields: birth_date, phone, email, ssn."""
    client = await client_portal_service.update_client_portal_info(
        db,
        client_id=client_id,
        user_id=current_user.id,
        birth_date=body.birth_date,
        phone=body.phone,
        email=body.email,
        ssn=getattr(body, "ssn", None),
    )
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    # Eagerly load accounts for full response
    await db.refresh(client, attribute_names=["accounts"])
    # Attach masked SSN
    from app.core.encryption import decrypt_ssn, mask_ssn
    if client.ssn_encrypted:
        try:
            client.ssn_masked = mask_ssn(decrypt_ssn(client.ssn_encrypted))
        except Exception:
            client.ssn_masked = None
    else:
        client.ssn_masked = None
    return client


@router.post("/migrate-account-types")
async def migrate_account_types(
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """pension1/pension2/pension_saving → pension 으로 일괄 변환."""
    mapping = {"pension1": "pension", "pension2": "pension", "pension_saving": "pension"}
    total = 0
    for old_val, new_val in mapping.items():
        result = await db.execute(
            sa_update(ClientAccount)
            .where(ClientAccount.client_id.in_(
                select(Client.id).where(Client.user_id == current_user.id)
            ))
            .where(ClientAccount.account_type == old_val)
            .values(account_type=new_val)
        )
        total += result.rowcount
    await db.commit()
    return {"migrated": total}
