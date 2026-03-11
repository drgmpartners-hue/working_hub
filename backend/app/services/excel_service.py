"""Excel parsing service using openpyxl."""
from __future__ import annotations

import io
from typing import Any


def parse_excel_bytes(content: bytes) -> dict[str, Any]:
    """Parse an Excel workbook from raw bytes.

    Reads every sheet in the workbook and converts each row (after the header
    row) into a dict keyed by the column headers found in row 1.

    Returns a dict with the following structure::

        {
            "sheets": ["Sheet1", "Sheet2", ...],
            "total_rows": 42,          # sum of data rows across all sheets
            "data": {
                "Sheet1": [{"col_a": val, ...}, ...],
                "Sheet2": [...],
            }
        }

    Raises ``ValueError`` if the bytes cannot be parsed as an xlsx workbook.
    """
    try:
        import openpyxl  # deferred import – always installed but keep coupling local
    except ImportError as exc:  # pragma: no cover
        raise ImportError("openpyxl is required for Excel parsing") from exc

    try:
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ValueError(f"Cannot parse Excel file: {exc}") from exc

    result: dict[str, list[dict[str, Any]]] = {}
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # First row is the header
        try:
            header_row = next(rows_iter)
        except StopIteration:
            result[sheet_name] = []
            continue

        headers: list[str] = [
            str(cell) if cell is not None else f"col_{i}"
            for i, cell in enumerate(header_row)
        ]

        sheet_data: list[dict[str, Any]] = []
        for row in rows_iter:
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue
            row_dict: dict[str, Any] = {}
            for header, cell in zip(headers, row):
                # Convert non-JSON-serialisable types (e.g. datetime) to str
                if hasattr(cell, "isoformat"):
                    row_dict[header] = cell.isoformat()
                else:
                    row_dict[header] = cell
            sheet_data.append(row_dict)

        result[sheet_name] = sheet_data
        total_rows += len(sheet_data)

    wb.close()

    return {
        "sheets": wb.sheetnames,
        "total_rows": total_rows,
        "data": result,
    }
